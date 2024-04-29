import telebot
from telebot import types
from aiogram import Router
from aiogram import types, F, Router
from aiogram.types import Message
from aiogram.filters import Command 
from aiogram.types import ContentType
import asyncio
import logging
from aiogram import Bot, Dispatcher
from aiogram.enums.parse_mode import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.client.bot import DefaultBotProperties
from aiogram.types import InputFile
import openpyxl
import apiinfo
import sqlite3
import os
import aiosqlite
from aiogram.types import FSInputFile, URLInputFile, BufferedInputFile
import csv



router = Router()
bot = Bot(token=apiinfo.API_KEY, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
dp.include_router(router)

@router.message(Command("start"))
async def start_handler(msg:Message):
    await msg.answer("Введите /info для просмотра справки доступных команд")


@dp.message(Command("info"))
async def info_handler(msg:Message):
    await  msg.answer("Доступные комманды:\n/edit_train - Позволяет отредактировать выбранную тренировку.\
                         \n/create_train - Создать тренировку.\
                         \n/choose_standart_train - Выбрать тренировку из списка стандартных.\
                         \n/import_your_train - Импортировать свою тренировку\
                         \n/send_sample_file - Бот отпрвит вам файл пример для импорта из XLSX или CSV")



@dp.message(Command("import_your_train"))
async def import_handler(msg:Message):
    kb = [
        [types.KeyboardButton(text="CSV")],
        [types.KeyboardButton(text="XLSX")]
    ]
    keyboard = types.ReplyKeyboardMarkup(keyboard=kb)
    await msg.answer("Выберите формат: CSV или XLSX.\
                         \n\nНажмите на одну из кнопок ниже.", reply_markup=keyboard)
    
@dp.message(lambda message: message.text == "CSV" or message.text == "XLSX")
async def format_selected(msg: Message):
    await msg.answer("Первые семь столбцов таблицы отвечают за дни недели. Ниже в каждой ячейке столбца указывайте упражнение.")
    if msg.text == "CSV":
        await msg.answer("Вы выбрали формат: CSV")
    elif msg.text == "XLSX":
        await msg.answer("Вы выбрали формат: XLSX. Пожалуйста, отправьте файл в этом формате.")


@dp.message(lambda message: message.document is not None)
async def handle_document(message: types.Message):
    await message.reply("Спасибо за отправку файла!")
    # Указываем более подходящую директорию для сохранения файлов
    # Например, папка 'downloads' в директории пользователя
    user_folder = os.path.expanduser('~')  # Получаем путь к домашней директории пользователя
    download_path = os.path.join(user_folder, 'downloads')

    # Проверяем, существует ли директория; если нет, создаем ее
    if not os.path.exists(download_path):
        os.makedirs(download_path)

    # Путь, куда будет сохранён файл
    destination = os.path.join(download_path, message.document.file_name)

    # Получаем объект файла
    file_info = await bot.get_file(message.document.file_id)

    # Скачиваем файл
    await bot.download_file(file_info.file_path, destination)
    file_extension = os.path.splitext(destination)[1]
    if file_extension.lower() == '.xlsx':
        print(0)
        data = read_xlsx(destination)

    elif file_extension.lower() == '.csv':
        print(1)
        data = read_csv(destination)
    else:
        await message.reply("Неподдерживаемый тип файла. Пожалуйста, отправьте файл в формате .xlsx или .csv.")
        return
    # Уведомляем пользователя о сохранении файла
    await message.reply(f"Файл был сохранен по пути: {destination}")
    user_id = message.from_user.id
    await save_to_db(user_id, data)


#Чтение данных из CSV
def read_csv(file_path):
    data = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)  # Пропускаем заголовок
        
        # Инициализация списка списков для столбцов
        for row in reader:
            # Расширяем список столбцов, если это необходимо
            if len(data) < len(row):
                for _ in range(len(row) - len(data)):
                    data.append([])
            # Добавляем данные в соответствующие списки столбцов
            for i, value in enumerate(row):
                if i < 7:  # Ограничиваем чтение только первыми семью колонками
                    data[i].append(value)

    # Форматирование данных для вывода
    # Преобразуем список списков в строку, где столбцы разделены двумя переносами строк
    formatted_data = '\n\n'.join(['\n'.join(col) for col in data])
    return formatted_data

# Чтение данных из xlsx файла
def read_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_col = sheet.max_column  # Получаем количество столбцов
    data = []

    # Собираем данные по столбцам
    for col in sheet.iter_cols(min_row=2, max_col=max_col, values_only=True):
        column_data = [cell if cell is not None else "" for cell in col]  # Замена None на пустую строку
        data.append(column_data)

    # Преобразуем список списков в строку, где столбцы разделены двумя переносами строк
    formatted_data = '\n\n'.join(['\n'.join(map(str, col)) for col in data])
    return formatted_data


async def save_to_db(user_id, data):
    async with aiosqlite.connect('kl1de.db') as db:
        # Создаем таблицу, если она не существует
        await db.execute('CREATE TABLE IF NOT EXISTS Trains (id INTEGER PRIMARY KEY, train TEXT)')

        # Преобразуем все строки входных данных в одну строку текста
        new_train_data = ''.join(''.join(str(item) for item in row) for row in data)
        new_train_data = '\n' + new_train_data
        # Проверяем наличие существующих данных для этого пользователя
        cursor = await db.execute('SELECT train FROM Trains WHERE id = ?', (user_id,))
        existing_train_data = await cursor.fetchone()

        if existing_train_data:
            # Если данные уже существуют, добавляем новые данные к существующим
            updated_train_data = existing_train_data[0] + new_train_data
            await db.execute('UPDATE Trains SET train = ? WHERE id = ?', (updated_train_data, user_id))
        else:
            # Добавляем новые данные, если для этого пользователя их еще нет
            await db.execute('INSERT INTO Trains (id, train) VALUES (?, ?)', (user_id, new_train_data))

        # Фиксируем изменения в базе данных
        await db.commit()




async def main():
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    asyncio.run(main())